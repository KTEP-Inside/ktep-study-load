import logging

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from ..models import *

logger = logging.getLogger(__name__)  # подключение логирования

is_paid = False


def create_group(ws: Worksheet):
    """Специальность, группа и курс"""
    global is_paid

    try:
        speciality = ' '.join(ws['A3'].value.split()[3:-1]).strip()  # специальность
        name_group = ' '.join(ws['A3'].value.split()[3:])  # полная группа
    except AttributeError:
        return 'Неверный шаблон'

    if ws['A3'].value.split()[-2][-1] == 'п':  # проверка на платную группу
        is_paid = True

    num_course = Course.objects.get(pk=int(ws.title[0]))
    speciality_obj, created = Speciality.objects.get_or_create(name=speciality)  # специальность
    group, created = SpecialityHasCourse.objects.get_or_create(course=num_course,
                                                               speciality=speciality_obj,  # группы
                                                               name_group=name_group,
                                                               is_paid=is_paid)
    return group


def create_table(ws: Worksheet,
                 subject: Cell,
                 teachers: Cell,
                 group: SpecialityHasCourse, ):
    """Препод и предмет, идём далее"""

    global is_paid

    if subject.value is not None:
        # ПП И ПДП - делим значение каждой ячейки на количество преподов
        if subject.value.title() in ('Всего', 'Итого'):
            is_paid = True  # Вне бюджетные записи

    if teachers.value is not None and subject.value is not None:  # предмет + препод

        teacher_mod = None
        if subject.value.strip().startswith('ПП') or subject.value.strip().startswith('ПДП')\
                or subject.value.strip() == 'Преддипломная практика':
            teacher_mod = len(teachers.value.split(','))

        if teacher_mod and not group.is_paid:
            subject_obj = Subject.objects.get_or_create(name=subject.value.strip(), is_paid=False)[0]
        # если это платный предмет в бесплатной группе
        elif is_paid and not group.is_paid:
            subject_obj = Subject.objects.get_or_create(name=subject.value.strip(), is_paid=is_paid)[0]
        # у платной группы свой флаг, поэтому все предметы платные ,и у них нет флага. Так же у бесплатной
        else:
            subject_obj = Subject.objects.get_or_create(name=subject.value.strip())[0]

        group_has_subject = GroupHasSubject.objects.get_or_create(group=group, subject=subject_obj)[0]

        for teacher_id, teacher in enumerate(teachers.value.split(',')):
            teacher_obj = Teacher.objects.get_or_create(name=teacher.strip())[0]

            create_type_load(ws, subject=subject,
                             group_has_subject=group_has_subject,
                             teacher_id=teacher_id,
                             teacher_obj=teacher_obj,
                             teacher_mod=teacher_mod)


def create_type_load(ws: Worksheet,
                     subject: Cell,
                     group_has_subject: GroupHasSubject,
                     teacher_id,
                     teacher_obj: Teacher,
                     teacher_mod):
    """Тип нагрузки и вызов создания записей по семестрам"""
    for load in ws.iter_cols(min_col=4, min_row=4, max_row=4):  # все типы нагрузки

        if load[0].value is not None:  # Если есть значение

            if load[0].value.find('Всего часов') == -1:  # завершаем по достижению
                type_load = TypeLoad.objects.filter(name=load[0].value.strip()).exists()

                if type_load:
                    type_load_obj = TypeLoad.objects.get(name=load[0].value.strip())

                    semester_load_writer(ws, load=load, subject=subject,
                                         type_load_obj=type_load_obj,
                                         group_has_subject=group_has_subject,
                                         teacher_obj=teacher_obj,
                                         teacher_id=teacher_id,
                                         teacher_mod=teacher_mod)
        else:

            continue


def semester_load_writer(ws: Worksheet,
                         load,
                         subject: Cell,
                         group_has_subject: GroupHasSubject,
                         type_load_obj: TypeLoad,
                         teacher_id,
                         teacher_obj: Teacher,
                         teacher_mod):
    """Валидация и итоговые записи"""

    semester = 1

    for cell in ws.iter_cols(min_col=load[0].column, max_col=load[0].column + 1,
                             min_row=subject.row, max_row=subject.row):
        semester_obj = Semester.objects.get(pk=semester)  # получаем номер семестра

        # возможно убрать, в шаблоне вроде не будет merge ячеек
        cur_cell = check_merge_cell(ws, cell, teacher_id)  # проверка на merge и взятие значение
        create_load(cur_cell, semester_obj=semester_obj, type_load_obj=type_load_obj,
                    group_has_subject=group_has_subject,
                    teacher=teacher_obj, teacher_mod=teacher_mod)
        semester += 1


def check_merge_cell(ws: Worksheet, cell, teacher_id) -> str | int | float:
    """Проверка на mergecell"""
    main_cell = False  # вынос в функцию

    for merged_range in ws.merged_cells.ranges:  # если это не главная ячейка
        if cell[0].coordinate in merged_range:
            main_cell = merged_range.min_row  # берем главную ячейку
            main_cell = ws.cell(main_cell, cell[0].column).value  # берем её значение

    if main_cell:  # если не главная ячейка
        cur_cell = main_cell
    else:
        cur_cell = check_multi_value(cell, teacher_id)
    return cur_cell


def check_multi_value(cell, teacher_id):
    if ',' in str(cell[0].value) and teacher_id <= len(cell[0].value.split(',')) - 1:
        cur_cell = cell[0].value.split(',')[teacher_id]
        if cur_cell.strip().isdigit():
            cur_cell = int(cur_cell)
    else:
        cur_cell = cell[0].value

    return cur_cell


def create_load(cur_cell: str | int | float,
                semester_obj: Semester,
                type_load_obj: TypeLoad,
                group_has_subject: GroupHasSubject,
                teacher: Teacher,
                teacher_mod
                ):
    """Создание часов нагрзуки на определённый семестр, кол-во часов и тп"""

    # Проверка на тип добавления
    if not isinstance(cur_cell, int) and cur_cell not in ['ДЗ', 'Э']:
        cur_cell = 0

    kwargs = {'semester': semester_obj,
              'type_load': type_load_obj,
              'group_has_subject': group_has_subject,
              'unallocated_hours': 0,
              }

    data = HoursLoad.objects.get_or_create(**kwargs)[0]
    if cur_cell in ['ДЗ', 'Э']:
        exam_obj = Exam.objects.get(exam=cur_cell)  # берем ДЗ или Э
        data.exam = exam_obj
        data.hours = 0
        data.save()
    else:
        if teacher_mod:
            data.hours = cur_cell // teacher_mod
        else:
            data.hours = cur_cell
        data.exam = None
        data.save()

    obj = TeacherHours.objects.get_or_create(teacher=teacher, hours_load=data)[0]
    obj.cur_hours = data.hours
    obj.cur_exam = data.exam
    obj.save()


def main_func(ws: Worksheet):
    """Основной скрипт получения данных с excel"""

    group = create_group(ws)

    # получаем преподавателей + их предмет
    for subject, teachers in ws.iter_rows(min_col=2, max_col=3, min_row=7):
        create_table(ws, subject=subject, teachers=teachers, group=group)


def load_data(file):
    global is_paid
    wb = openpyxl.load_workbook(file)  # открываем файл
    all_sheets = wb.sheetnames  # все листы
    for sheet_id, _ in enumerate(all_sheets):  # проходим по всем листам
        ws = wb.worksheets[sheet_id]
        main_func(ws)
        is_paid = False
