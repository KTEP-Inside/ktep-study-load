import openpyxl
from bs4 import BeautifulSoup
from string import ascii_uppercase
from openpyxl.styles import Font, Alignment, Border, Side


class SetCellStyles:
    """Класс для удобной установки стилей ячейки"""

    common_font = Font(name='Times New Roman', size=14, bold=True)
    common_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    common_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                           bottom=Side(style='thin'))
    no_border_style = Border(left=None, right=None)

    def __init__(self, cell):
        self.cell = cell

    def set_styles(self):
        self.cell.font = self.common_font
        self.cell.alignment = self.common_alignment
        self.cell.border = self.common_border

    def delete_borders(self):
        self.cell.border = self.no_border_style


def change_cells_height(ws) -> None:
    """Меняем высоту строк под шаблон"""
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 30
    total_rows = ws.max_row
    ws.row_dimensions[total_rows].height = 25
    ws.row_dimensions[total_rows - 1].height = 25
    ws.row_dimensions[total_rows - 2].height = 25


def change_cells_width(ws) -> None:
    """Меняем ширину колонок под шаблон"""
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['P'].width = 9
    ws.column_dimensions['Q'].width = 9
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 10


def create_table_styles(ws):
    """Стилизуем шаблон"""

    # Проход по всем ячейкам и установка общего шрифта
    for row in ws.iter_rows():
        for cell in row:
            cell = SetCellStyles(cell)
            cell.set_styles()

    change_cells_height(ws)
    change_cells_width(ws)

    # тут мы меняем стиль первой строки таблицы
    ws.merge_cells('J1:X1')
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=23):
        for cell in row:
            cell = SetCellStyles(cell)
            cell.delete_borders()


def get_html_table_header(soup):
    """Получаем данные из header-a таблицы"""
    teacher_load = soup.find('td', 'teacher-load').text.split('\n', 1)[0]
    teacher = soup.find('select', class_='teacher').find('option').string
    year = soup.find('div', class_='education_year').string

    return teacher_load, teacher, year


def get_semesters_data(types_load_elements) -> list:
    """Получаем список семестров"""
    semesters = []
    for i in range(1, (len(types_load_elements)) * 2 + 1):
        if i % 2 == 0:
            semesters.append('II')
        else:
            semesters.append('I')
    return semesters


def get_hours_load_data(soup, types_load_elements, hours_load_tr) -> list:
    """Получаем часы нагрузки"""
    hours_load_elements = []
    last_row_in_hours = 0

    for row in range(1, len(hours_load_tr) + 1):
        lst = []
        group = soup.find(id=f'group_{row}').find('option')
        subject = soup.find(id=f'subject_{row}').find('option')

        if group.string is not None and subject is not None:
            last_row_in_hours += 1

            lst.append(last_row_in_hours)
            lst.append(subject.string)
            lst.append(group.string)

            for type_l in range(1, len(types_load_elements) + 1):
                lst.append(soup.find(id=f'type-load_{row}_{type_l}_1')['value'])
                lst.append(soup.find(id=f'type-load_{row}_{type_l}_2')['value'])

            hours_load_elements.append(lst)
            lst.append(soup.find(id=f'budget_0_{row}')['value'])
            lst.append(soup.find(id=f'budget_1_{row}')['value'])
            lst.append(soup.find(id=f'budget_2_{row}')['value'])

    return hours_load_elements


def get_results_budget(soup, results) -> list:
    """Получаем результаты подсчета часов нагрузки"""
    results_elements = []

    for row in range(1, len(results) + 1):
        lst = [int(soup.find(id=f'budget_sum_{row}')['value']),
               int(soup.find(id=f'extra_budget_sum_{row}')['value']),
               int(soup.find(id=f'budget_result_{row}')['value'])]
        results_elements.append(lst)

    return results_elements


def get_html_data(soup):
    """Получаем данные из основной части таблицы"""
    hours_load_tr = soup.find('tbody').find_all('tr')
    results = soup.find('tfoot').find_all('tr')

    types_load_elements = [elem.text for elem in soup.find_all('td', class_='type_load')]
    info_elements = [elem.text for elem in soup.find_all('td', class_='info')]
    budget_elements = [elem.text for elem in soup.find_all('td', class_='budget')]
    results_info = [i.string for i in soup.find_all('td', class_='budget-info')]

    results_elements = get_results_budget(soup, results)
    hours_load_elements = get_hours_load_data(soup, types_load_elements, hours_load_tr)
    semesters = get_semesters_data(types_load_elements)

    return types_load_elements, info_elements, budget_elements, results_info,\
        results_elements, hours_load_elements, semesters


def create_header(soup, ws) -> None:
    """Создаем первую строку таблицы в excel"""
    teacher_load, teacher, year = get_html_table_header(soup)

    header = [teacher_load, teacher, year]
    header_length = 0
    for i in range(len(header)):
        ws.cell(row=1, column=1 + header_length, value=header[i])
        ws.merge_cells(f'{ascii_uppercase[header_length]}1:{ascii_uppercase[header_length + 2]}1')
        header_length += 3


def create_table_cols(ws, info_elements,
                      types_load_elements,
                      budget_elements,
                      semesters) -> None:
    """Заполняем колонки названиями"""
    last_row = ws.max_row + 1

    for col, info in enumerate(info_elements):
        ws.cell(row=last_row, column=col+1, value=info)
        ws.merge_cells(f'{ascii_uppercase[col]}{last_row}:{ascii_uppercase[col]}{last_row + 1}')

    for col, header in zip(range(4, len(types_load_elements) * 2 + 5, 2), types_load_elements):
        ws.cell(row=last_row, column=col, value=header)
        ws.merge_cells(f'{ascii_uppercase[col - 1]}{last_row}:{ascii_uppercase[col]}{last_row}')

    last_col = ws.max_column + 1
    for col, budget_info in enumerate(budget_elements):
        ws.cell(row=last_row, column=last_col+col, value=budget_info)
        ws.merge_cells(f'{ascii_uppercase[last_col+col-1]}{last_row}:{ascii_uppercase[last_col+col-1]}{last_row + 1}')

    last_row = ws.max_row
    for col, semester in enumerate(semesters):
        ws.cell(row=last_row, column=col+4, value=semester)


def create_table_rows(ws, hours_load_elements,
                      results_info,
                      results_elements) -> None:
    """заполняем строки часами нагрузки"""
    for row in hours_load_elements:
        ws.append(row)
        total_rows = ws.max_row
        ws.row_dimensions[total_rows].height = 40

    for res in range(len(results_info)):
        ws.append([results_info[res]])
        ws.merge_cells(f'{ascii_uppercase[1]}{ws.max_row}:{ascii_uppercase[ws.max_column-4]}{ws.max_row}')
        for i in range(len(results_elements)):
            ws.cell(row=ws.max_row, column=ws.max_column-2+i, value=results_elements[res][i])


def create_excel_report_data(data):
    """Основная функция, создает excel файл и возвращает его"""

    soup = BeautifulSoup(data, 'html.parser')
    teacher_load, teacher, year = get_html_table_header(soup)
    types_load_elements, info_elements, budget_elements, results_info, \
        results_elements, hours_load_elements, semesters = get_html_data(soup)

    wb = openpyxl.Workbook()
    ws = wb.active

    create_header(soup, ws)
    create_table_cols(ws, info_elements, types_load_elements, budget_elements, semesters)
    create_table_rows(ws, hours_load_elements, results_info, results_elements)

    create_table_styles(ws)

    filename = f'{teacher}-отчет-{year}'
    return wb, filename

